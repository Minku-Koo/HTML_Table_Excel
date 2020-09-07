#!/usr/bin/python
# -*- coding: utf-8 -*-

'''
- Scrapping HTML Table and Input a Table Data to Excel
- This Library apply HTML_Table_Extractor

- Library Name : Table_Excel
- Created Date : 27/Aug/2020
- Author : Minku Koo
- E-Mail : corleone@kakao.com
- Version : 1.0
- Keywords : 'Excel', 'Table', 'HTML', 'Crawling', 'Selenium', 'Extractor'
- Github URL : https://github.com/Minku-Koo/HTML_Table_Excel

 * How to Use?
TableExcel = Table_Excel( URL_list <type=(String)list>, ChromeDriver Path <type=String>)
TableExcel.makeExel_abs( Excel File Path <type=String>, Table Header Color by Hex <type=String> (Default=F8E0EC) )
TableExcel.makeExel_sep( Excel File Path <type=String> )

 * Please, Import these Library : BeautifulSoup4, openpyxl, time, selenium
 * You should check your ChromeDriver exist
 * Also, You have to check, that your Chrome Version and your ChromeDriver version is same
'''

# for Extractor class
from bs4 import BeautifulSoup, Tag
import os
import csv
import pdb

# for Table_Excel class
from openpyxl.styles import PatternFill, Color
import openpyxl
from openpyxl.styles.borders import Border, Side

import time
import selenium
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait

# HTML_Table_Extractor  Transform
class Extractor(object): 
    def __init__(self, input, id_=None, **kwargs):
        # TODO: should divide this class into two subclasses
        # to deal with string and bs4.Tag separately

        # validate the input
        if not isinstance(input, str) and not isinstance(input, Tag):
            raise Exception('Unrecognized type. Valid input: str, bs4.element.Tag')

        soup = BeautifulSoup(input, 'html.parser').find() if isinstance(input, str) else input

        # locate the target table
        if soup.name == 'table':
            self._table = soup
        else:
            self._table = soup.find(id=id_)

        if 'transformer' in kwargs:
            self._transformer = kwargs['transformer']
        else:
            self._transformer = str

        self._output = []
        self._cellinfo=[]
        self._isHeader=[]

    def parse(self):
        self._output = []
        row_ind = 0
        col_ind = 0
        # for row in self._table.find_all('tr'):
        for tab in self._table.find_all("table"): # 엑셀에서 테이블 간격 주기 위함 - 한 줄 띄움
            try:
                for i in tab.find('tr'):
                    self._insert(row_ind, col_ind, 0, 0, "",0) # 엑셀 테이블 사이 간격 한 줄
                row_ind += 1
            except UnicodeEncodeError:
                raise Exception( 'Failed to decode text; you might want to specify kwargs transformer=unicode' )
            for row in tab.find_all('tr'):
            
                # record the smallest row_span, so that we know how many rows
                # we should skip
                smallest_row_span = 1

                for cell in row.children:
                    if cell.name in ('td', 'th'):
                        # check multiple rows
                        # pdb.set_trace()
                        row_span = int(cell.get('rowspan')) if cell.get('rowspan') else 1

                        # try updating smallest_row_span
                        smallest_row_span = min(smallest_row_span, row_span)

                        # check multiple columns
                        col_span = int(cell.get('colspan')) if cell.get('colspan') else 1

                        # find the right index
                        while True:
                            if self._check_cell_validity(row_ind, col_ind):
                                break
                            col_ind += 1

                        # insert into self._output
                        cell_value = cell.get_text()
                        isHeader =0
                        if cell.name =="th": # 테이블 헤더일 경우 - 셀 컬러 변경 위함
                            isHeader = 1
                        
                        try:
                            self._insert(row_ind, col_ind, row_span, col_span, self._transformer( cell_value ),isHeader)
                        except UnicodeEncodeError:
                            raise Exception( 'Failed to decode text; you might want to specify kwargs transformer=unicode' )

                        # update col_ind
                        col_ind += col_span

                # update row_ind
                row_ind += smallest_row_span
                col_ind = 0
        return self

    def return_list(self): #결과 리스트 출력
        return self._output
    
    def return_cellinfo(self): # 병합해야할 테이블 셀 정보 리스트 출력
        return self._cellinfo
    
    def return_isHeader(self): # 헤더 유무 판별하는 리스트 출력
        return self._isHeader
        
    def write_to_csv(self, path='.', filename='output.csv'): # 테이블을 csv 파일로 생성
        with open(os.path.join(path, filename), 'w', encoding="utf-8") as csv_file:
            table_writer = csv.writer(csv_file)
            for row in self._output:
                input_row = [r.replace("\n","").replace("\t","") for r in row] #개행, 탭 제거
                table_writer.writerow(input_row)
        return

    def _check_validity(self, i, j, height, width):
        """
        check if a rectangle (i, j, height, width) can be put into self.output
        """
        return all(self._check_cell_validity(ii, jj) for ii in range(i, i+height) for jj in range(j, j+width))

    def _check_cell_validity(self, i, j):
        """
        check if a cell (i, j) can be put into self._output
        """
        if i >= len(self._output):
            return True
        if j >= len(self._output[i]):
            return True
        if self._output[i][j] is None:
            return True
        return False

    def _insert(self, i, j, height, width, val, isHeader):
        # pdb.set_trace()
        info =""
        if height!=1: #rowspan일 경우
            info  = "r"+str(height)
        if width!=1: #colspan일 경우
            info  = "c"+str(width)
        
        
        for ii in range(i, i+height): #테이블 크기만큼 값 채워주기
            for jj in range(j, j+width):
                self._insert_cell(ii, jj, val)
                if info !="": # 병합되야하는 경우
                    self._cellinfo.append( ( ii+1,jj+1, info))
                if isHeader ==1: #헤더일 경우
                    self._isHeader.append(( ii+1,jj+1) )

    def _insert_cell(self, i, j, val): # output에 값 채워넣기
        while i >= len(self._output):
            self._output.append([])
        while j >= len(self._output[i]):
            self._output[i].append(None)

        if self._output[i][j] is None:
            self._output[i][j] = val


class Table_Excel(object):
    def __init__(self, url_list, path="./chromedriver.exe"): # URL 리스트, 크롬 드라이버 경로
        self.url_list = url_list
        self.thin_border = Border(left=Side(style='thin'), #헤더 테두리 굵게 표시하기 위함
                        right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        self.headers = {'User-Agent' : 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) '}
        
        options = webdriver.ChromeOptions() #크름 드라이버를 보이지 않게 하기 위한 option 추가
        options.add_argument('headless')
        options.add_argument('--disable-gpu')
        options.add_argument('lang=ko_KR')
        
        self.driver = webdriver.Chrome(path, chrome_options=options) #크롬 드라이버 생성
        self.driver.implicitly_wait(5)
    
    def table_parse(self, url): # URL을 통해 테이블 정보, 병합 셀 정보, 제목, 헤더 여부를 판단
        self.driver.get(url)
        self.driver.implicitly_wait(10)
        time.sleep(0.5)
        self.soup = BeautifulSoup(self.driver.page_source, 'html.parser')
        
        ext = Extractor(self.soup) # Extractor 이용
        ext.parse()
        data =ext.return_list() # 테이블 데이터를 리스트로 받음
        absorption =ext.return_cellinfo() # 테이블 병합 셀 정보를 리스트로 받음
        isHeader =ext.return_isHeader() # 테이블 헤더 여부를 리스트로 받음
        title = self.soup.find("title") # 제목을 추출
        if title == None or title.get_text()== "": #제목이 없을 경우
            title = 'no title'
        else: title = title.get_text()
        
        return data, absorption, title, isHeader
    
    def cell_merge(self, pos, absorption, isHeader): #셀 위치/ 셀 병합 정보/ 셀 헤더 여부 판단
        row = pos.row # 수직
        col = pos.column # 수평
        
        cell_abs = [ a[:-1] for a in absorption] #병합 셀 위치만 리스트로 변환
        cell_way = [ a[-1] for a in absorption] # 병합할 방향과 셀 개수를 리스트로 변환
        abs_cell_count = None # 병합해야할 셀의 개수 
        rows, cols = 0,0 # 현재 셀에서 병합해야할 목적이 되는 셀
        
        # row에서 2를 빼는 이유 : 엑셀 1, 2 라인에 URL, TITLE 정보를 입력했기 때문
        if (row-2, col) in cell_abs: # rowspan or colspan이 있는 경우
            
            indx = cell_abs.index( (row-2, col) )
            where = cell_way[indx][0] #병합할 셀 위치
            abs_cell_count = cell_way[indx][1] #병합할 방향과 개수
            if where=="r": # rowspan일 경우
                rows = row+int(abs_cell_count)-1
                cols = col
            else: # colspan일 경우
                cols = col+int(abs_cell_count)-1
                rows = row
            
        head =0
        if (row-2, col) in isHeader: # 해당 셀이 헤더일 경우 
            head =1
            
        return abs_cell_count, head, row, col, rows, cols
    
    
    def makeExel_sep(self, filename='makeExel_sep.xlsx'): # 분할된 엑셀 테이블 생성
        wb_sep = openpyxl.Workbook()
        a=True
        for url in self.url_list:
            data, absorption, title, isHeader = self.table_parse(url)
            
            if a: #첫 링크일 경우
                sheet = wb_sep.active
                a=False
            else: sheet = wb_sep.create_sheet() # 두번째 이후 링크
        
            sheet.append(['URL',url]) # 시트 첫째줄에 링크
            sheet.append(['TITLE',title]) # 시트 둘째줄에 제목
            for i in data: #셋째줄부터 아래로 테이블 정보 입력
                sheet.append(i)
        wb_sep.save(filename)
        
    def makeExel_abs(self, filename='makeExel_abs.xlsx', color='F8E0EC'):
        wb_abs = openpyxl.Workbook() 
        a=True
        for url in self.url_list:
            data, absorption, title, isHeader = self.table_parse(url)
            if a: #첫 링크일 경우
                sheet = wb_abs.active
                a=False
            else: sheet = wb_abs.create_sheet() # 두번째 이후 링크
            sheet.append(["URL",url]) # 시트 첫째줄에 링크
            sheet.append(["TITLE",title]) # 시트 둘째줄에 제목
            for i in data: #셋째줄부터 아래로 테이블 정보 입력
                sheet.append(i)
                
            for line in sheet: #모든 셀을 비교하며 병합 또는 헤더 판단
                for cel in line:
                    delcount, head, row, col, rows, cols = self.cell_merge(cel, absorption, isHeader)
                    if head ==1: #헤더일 경우
                        # cell color 변경 
                        sheet.cell(row=row, column=col).fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                        # cell border 변경
                        sheet.cell(row=cel.row, column = cel.column).border = self.thin_border
                    if delcount == None: #병합할 셀이 없는 경우
                        continue
                    #셀 병합
                    sheet.merge_cells(start_row=row, start_column=col, end_row=rows, end_column=cols)
                    # 병합한 셀 개수 만큼, 병합될 셀 리스트에서 삭제
                    for i in range(int(delcount)):
                        absorption.pop(0)
                    
        wb_abs.save(filename) # 엑셀 저장









