# -*- coding: utf-8 -*-

'''
- Extract HTML Table and Input a Table Data to Excel
- This Library apply HTML_Table_Extractor

- Library Name : Table_Excel
- Created Date : 27/Aug/2020
- Updated Date : 11/Mar/2021
- Author : Minku Koo
- E-Mail : corleone@kakao.com
- Version : 1.1.4
- Keywords : 'Excel', 'Table', 'HTML', 'Crawling', 'Selenium', 'Extractor'
- Github URL : https://github.com/Minku-Koo/HTML_Table_Excel

 * How to Use?
TableExcel = Table_Excel( URL_list <(String)list>, ChromeDriver Path <String>)
TableExcel.makeExel_abs( Excel File Path <String>, Table Header Color by Hex <String> (Default=F8E0EC) )
TableExcel.makeExel_sep( Excel File Path <String> )

 * You should check your ChromeDriver version
 * Also, You have to check, that your Chrome Browser Version and your ChromeDriver version is same
'''

# for Extractor class
from bs4 import BeautifulSoup, Tag
import os, csv

# for Table_Excel class
from openpyxl.styles import PatternFill, Color
from openpyxl.styles.borders import Border, Side
import openpyxl

import time, itertools
import selenium
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait


# HTML_Table_Extractor  Transform
class Extractor(object): 
    def __init__(self, input, id_=None, **kwargs):
        # TODO: should divide this class into two subclasses
        # to deal with string and bs4.Tag separately
        
        # validate the input
        if not isinstance(input, str) and not isinstance(input, Tag):
            raise Exception('Unrecognized type. Valid input: str, bs4.element.Tag')
        
        soup = BeautifulSoup(input, 'html.parser').find() if isinstance(input, str) else input
        
        # locate the target table
        if soup.name == 'table': self._table = soup
        else: self._table = soup.find(id=id_)

        if 'transformer' in kwargs:
            self._transformer = kwargs['transformer']
        else:
            self._transformer = str

        self._output = [] # 결과 저장 리스트
        self._cellinfo=[] # 병합 여부 판단
        self._isHeader=[] # th 태그 여부 판단
        self._xpath="" # xpath 임시 저장
        self.table_count = 0 # 중첩된 테이블인지의 여부 판단

    def xpath_soup(self, element): # HTML 요소의 xpath를 구하는 함수
        """
        Generate xpath of soup element
        :param element: bs4 text or node
        :return: xpath as string
        """
        components = []
        child = element if element.name else element.parent
        for parent in child.parents:
            """
            @type parent: bs4.element.Tag
            """
            previous = itertools.islice(parent.children, 0,parent.contents.index(child))
            xpath_tag = child.name
            xpath_index = sum(1 for i in previous if i.name == xpath_tag) + 1
            components.append(xpath_tag if xpath_index == 1 else '%s[%d]' % (xpath_tag, xpath_index))
            child = parent
        components.reverse()
        components.pop()
        
        return '/%s' % '/'.join(components)
    
    def parse(self):
        self._output = []
        row_ind, col_ind = 0, 0 # 테이블 cell 저장 위치 인덱스
        tr_count, td_count =0, 0 # 테이블 중첩 혹은 가로 배열 시, 조정을 위한 값
        td_count_list = [] #테이블 가로값 저장 for 가로 배열 테이블의 조정
        reit = 1 # 중첩 단계
        row_reit = 0 # 중첩된 테이블 개수를 판단하여,, 엑셀에 row를 조정하기 위한 계산
        innerTable = [] # 중첩된 테이블 리스트
        indiv_table = 0 # 중첩되지 않고, 가장 바깥쪽 테이블 count
        
        
        for tab in self._table.find_all("table"): # 엑셀에서 테이블 간격 주기 위함 - 한 줄 띄움
            
            tag_path = self.xpath_soup(tab) # xpath 활용해서 부모 태그 찾기
            isSameParents = False # 같은 부모 태그 가진지에 대한 여부
            if tag_path == self._xpath: # 공통된 부모 태그를 가질 경우
                isSameParents = True
            else: self._xpath = tag_path #독립된 부모 태그일 경우
            
            #가로 정렬 테이블인지 판단
            temp_row = 0
            if isSameParents:
                temp_row = len(tab.find_all('tr')) # calc_temp_row(tab)
                
                if tr_count != temp_row: # 테이블 세로값이 다를 경우, 가로 정렬 아님
                    isSameParents = False
                    
                # 이전 테그가 테이블일 경우, 가로 정렬 테이블로 판단
                prev_table = tab.find_previous_sibling() #이전 태그 추출
                #이전 태그가 없거나, 테이블 태그가 아니면 가로정렬 아닌걸로 판단
                if prev_table == None or prev_table.name != "table":
                    isSameParents = False
            
            reit = tag_path.count('table')+1 # xpath를 통해 테이블이 중첩인지 판단, 중첩 단계 판별
            if tag_path.count('table') == 0: # 테이블이 중첩되지 않은 테이블일 경우, 값 초기화
                reit =1
                row_ind -= row_reit
                row_reit=0
                indiv_table += 1 # 개별 테이블의 번호는 높여줌
                self.table_count=0
                
            try:
                if isSameParents == False: #같은 부모 아닐 경우
                    td_count =0 # 조정 값 초기화
                    row_ind += 1 # 한 줄 띄움
                for tr in tab.find_all('tr'):
                    td_count_list.append(( len(tr.find_all('td')) + len(tr.find_all('th')) ) ) #최초 테이블 가로값 추가
               
            except UnicodeEncodeError:
                raise Exception( 'Failed to decode text; you might want to specify kwargs transformer=unicode' )
                
            
            if len(innerTable) >0: # 중첩 테이블이 하나 이상이고, 새로운 테이블을 출력할 경우
                self._insert(row_ind-row_reit, col_ind, 1, 1, innerTable.pop(0), 0) #테이블 번호 삽입
                row_ind+=1
                
            
            for row in tab.find_all('tr'):
                tr_count = len(tab.find_all('tr')) #테이블 세로 값 판단
                
                if self.xpath_soup(row).count('table') == 0: #중첩되지 않은 테이블일 경우
                    reit =1 # 중첩 단계 초기화
                
                # record the smallest row_span, so that we know how many rows
                # we should skip
                smallest_row_span = 1
                innerTableIndex=0
                for cell in row.children:
                    if cell.name in ('td', 'th'):
                        
                        # check multiple rows
                        # pdb.set_trace()
                        row_span = int(cell.get('rowspan')) if cell.get('rowspan') else 1

                        # try updating smallest_row_span
                        smallest_row_span = min(smallest_row_span, row_span)

                        # check multiple columns
                        col_span = int(cell.get('colspan')) if cell.get('colspan') else 1

                        # find the right index
                        while True:
                            if self._check_cell_validity(row_ind, col_ind):
                                break
                            col_ind += 1
                        
                        cell_value = cell.get_text()
                        table_numb=0 # 중첩 테이블 번호 초기화
                        
                        if cell.find("table") != None: # td안에 table이 있을 경우
                            
                            for tables in cell.find_all('table'): # 셀 안에 테이블이 있을 경우
                                cell_value=cell_value.replace(tables.get_text(), "") # 기존 테이블 안의 내용은 삭제
                                self.table_count+=1 # 중첩 테이블 총 번호 증가
                                
                                if self.xpath_soup(cell).count('table') >1: # 셀이 테이블 중첩일 경우
                                    cell_value = innerTable[table_numb+innerTableIndex] + cell_value 
                                    self.table_count -=1
                                    # table_numb +=1 
                                    innerTableIndex+=1
                                    
                                elif self.xpath_soup(cell).count('table') == reit: #셀이 중첩 단계와 같을 경우
                                    table_name = ' [ TABLE '+str(indiv_table)+'-'+str(self.table_count)+' ] '
                                    # 셀 값에 테이블 번호 추가
                                    cell_value = table_name + cell_value
                                    # 중첩 테이블 리스트에 추가
                                    innerTable.append( table_name )
                                    table_numb +=1
                                else: 
                                    self.table_count -=1
                                    
                        if self.xpath_soup(cell).count('table') != reit: # 해당 셀이 중첩 단계와 다를 경우, 저장하지 않음
                            continue
                        
                        isHeader = 1 if cell.name =="th" else 0
                        
                        try:
                            td_counts, tr_counts = 0, 0
                            if isSameParents :
                                td_counts = td_count
                                tr_counts = tr_count
                                
                            # 값 최종 저장
                            self._insert(
                                            row_ind-row_reit, 
                                            col_ind, 
                                            row_span, 
                                            col_span, 
                                            self._transformer( cell_value ), 
                                            isHeader, 
                                            (tr_counts,td_counts)
                                        )
                        except UnicodeEncodeError:
                            raise Exception( 'Failed to decode text; you might want to specify kwargs transformer=unicode' )

                        # update col_ind
                        col_ind += col_span

                # update row_ind
                row_ind += smallest_row_span
                
                col_ind = 0
                row_table = row.find_all('table') # tr에 테이블 모두 가져옴
                if row_table== None: # 없을 경우
                    continue
                for tb in row_table: # 엑셀에서 row 조정을 위해 값 계산 과정
                    for _ in tb.find_all("tr"): # 테이블 있을 경우
                        if self.xpath_soup(tb).count('table') > reit: #해당 셀이 중첩 단계보다 높은 경우
                            break
                        row_reit+=1 
                
            
            td_count += max(td_count_list)
            td_count_list=[]
            
            if isSameParents: # 부모 태그가 같을 경우, 조정시키기 위함
                for _ in range(tr_count):
                    lastest  = self._output.pop() # 마지막 값을 가져와서 
                    self._output[len(self._output)-tr_count].extend(lastest) #이전 리스트에 추가
                row_ind -= tr_count # 엑셀 row 조절
                
            # 중첩이면 단계 1 추가
            if tab.find("table") != None: reit +=1
        
        return self

    def get_list(self): #결과 리스트 출력
        return self._output
    
    def get_cellinfo(self): # 병합해야할 테이블 셀 정보 리스트 출력
        return self._cellinfo
    
    def get_isHeader(self): # 헤더 유무 판별하는 리스트 출력
        return self._isHeader
        
    def write_to_csv(self, path='.', filename='output.csv'): # 테이블을 csv 파일로 생성
        with open(os.path.join(path, filename), 'w', encoding="utf-8") as csv_file:
            table_writer = csv.writer(csv_file)
            for row in self._output:
                input_row = [r.replace("\n","").replace("\t","") for r in row] #개행, 탭 제거
                table_writer.writerow(input_row)
        return 0

    def _check_validity(self, i, j, height, width):
        """
        check if a rectangle (i, j, height, width) can be put into self.output
        """
        return all(self._check_cell_validity(ii, jj) for ii in range(i, i+height) for jj in range(j, j+width))

    def _check_cell_validity(self, i, j):
        """
        check if a cell (i, j) can be put into self._output
        """
        if i >= len(self._output):
            return True
        if j >= len(self._output[i]):
            return True
        if self._output[i][j] is None:
            return True
        return False

    def _insert(self, i, j, height, width, val, isHeader, update=(0,0)):
        
        info = ""
        if height!=1: #rowspan일 경우
            info  = "r"+str(height)
        if width!=1: #colspan일 경우
            info  = "c"+str(width)
        
        for ii in range(i, i+height): #테이블 크기만큼 값 채워주기
            for jj in range(j, j+width):
                self._insert_cell(ii, jj, val)
                if info != "": # 병합되야하는 경우
                    self._cellinfo.append( ( ii+1-update[0], jj+1+update[1], info))
                if isHeader ==1: #헤더일 경우
                    self._isHeader.append( ( ii+1-update[0], jj+1+update[1]) )
                    
        return 0

    def _insert_cell(self, i, j, val): # output에 값 채워넣기
        
        while i >= len(self._output):
            self._output.append([])
        while j >= len(self._output[i]):
            self._output[i].append(None)
        
        if self._output[i][j] is None:
            self._output[i][j] = val
            
        return 0


class Table_Excel(object):
    def __init__(self, url_list, path="./chromedriver.exe"): # URL 리스트, 크롬 드라이버 경로
        self.url_list = url_list
        self.thin_border = Border(
                                    left=Side(style='thin'), #헤더 테두리 굵게 표시하기 위함
                                    right=Side(style='thin'), 
                                    top=Side(style='thin'), 
                                    bottom=Side(style='thin')
                                )
        
        self.headers = {'User-Agent' : 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) '}
        
        options = webdriver.ChromeOptions() #크름 드라이버를 보이지 않게 하기 위한 option 추가
        options.add_argument('headless')
        options.add_argument('--disable-gpu')
        options.add_argument('lang=ko_KR')
        
        self.driver = webdriver.Chrome(path, chrome_options=options) #크롬 드라이버 생성
        self.driver.implicitly_wait(5)
    
    def table_parse(self, url): # URL을 통해 테이블 정보, 병합 셀 정보, 제목, 헤더 여부를 판단
        self.driver.get(url)
        self.driver.implicitly_wait(10)
        time.sleep(0.2)
        
        self.soup = BeautifulSoup(self.driver.page_source, 'html.parser')
        ext = Extractor(self.soup) # Extractor 이용
        ext.parse()
        
        data =ext.get_list() # 테이블 데이터를 리스트로 받음
        absorption =ext.get_cellinfo() # 테이블 병합 셀 정보를 리스트로 받음
        isHeader =ext.get_isHeader() # 테이블 헤더 여부를 리스트로 받음
        title = self.soup.find("title") # 제목을 추출
        
        if title == None or str(type(title)) != "<class 'bs4.element.Tag'>": #제목이 없을 경우
            title = 'no title'
        elif title.get_text()== "": #제목이 없을 경우
            title = 'no title'
        else: title = title.get_text()
        
        return data, absorption, title, isHeader
    
    def cell_merge(self, position, absorption, isHeader): #셀 위치/ 셀 병합 정보/ 셀 헤더 여부 판단
        row = position.row # 수직
        col = position.column # 수평
        
        cell_abs = [ a[:-1] for a in absorption ] #병합 셀 위치만 리스트로 변환
        cell_direction = [ a[-1] for a in absorption ] # 병합할 방향과 셀 개수를 리스트로 변환
        abs_cell_count = None # 병합해야할 셀의 개수 
        rows, cols = 0,0 # 현재 셀에서 병합해야할 목적이 되는 셀
        
        # row에서 2를 빼는 이유 : 엑셀 1, 2 라인에 URL, TITLE 정보를 입력했기 때문
        if (row-2, col) in cell_abs: # rowspan or colspan이 있는 경우
            
            index = cell_abs.index( (row-2, col) )
            direction = cell_direction[index][0] #병합할 셀 위치
            abs_cell_count = int( cell_direction[index][1:] ) #병합할 방향과 개수
            
            if direction=="r": # rowspan일 경우
                rows = row+abs_cell_count-1
                cols = col
            else: # colspan일 경우
                cols = col+abs_cell_count-1
                rows = row
        
        head = 1 if (row-2, col) in isHeader else 0
        
        return abs_cell_count, head, row, col, rows, cols
    
    def makeExel_sep(self, filename='makeExel_sep.xlsx'): # 분할된 엑셀 테이블 생성
        excel_sep = openpyxl.Workbook()
        isFirstPage=True
        for url in self.url_list:
            data, _ , title, _ = self.table_parse(url)
            
            sheet = excel_sep.active if isFirstPage else excel_sep.create_sheet()
            isFirstPage=False
            
            sheet.append(['URL',url]) # 시트 첫째줄에 링크
            sheet.append(['TITLE',title]) # 시트 둘째줄에 제목
            for i in data: #셋째줄부터 아래로 테이블 정보 입력
                sheet.append(i)
                
        excel_sep.save(filename)
        
        return 0
        
    def makeExel_abs(self, filename='makeExel_abs.xlsx', color='F8E0EC'):
        excel_abs = openpyxl.Workbook() 
        
        isFirstPage=True
        for url in self.url_list:
            data, absorption, title, isHeader = self.table_parse(url)
            
            sheet = excel_abs.active if isFirstPage else excel_abs.create_sheet()
            isFirstPage=False
            
            sheet.append(["URL",url]) # 시트 첫째줄에 링크
            sheet.append(["TITLE",title]) # 시트 둘째줄에 제목
            for i in data: #셋째줄부터 아래로 테이블 정보 입력
                sheet.append(i)
                
            for line in sheet: #모든 셀을 비교하며 병합 또는 헤더 판단
                for cell in line:
                    delcount, head, row, col, rows, cols = self.cell_merge(cell, absorption, isHeader)
                    if head ==1: #헤더일 경우
                        # cell color 변경 
                        sheet.cell(row=row, column=col).fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                        # cell border 변경
                        sheet.cell(row=cell.row, column = cell.column).border = self.thin_border
                    
                    if delcount == None: #병합할 셀이 없는 경우
                        continue
                    #셀 병합
                    sheet.merge_cells(start_row=row, start_column=col, end_row=rows, end_column=cols)
                    # 병합한 셀 개수 만큼, 병합될 셀 리스트에서 삭제
                    if row == rows: # col
                        for a in range(cols-col+1):
                            absorption.remove( (row-2, col+a, "c"+str(cols-col+1)) )
                    if col == cols: #row
                        for a in range(rows-row+1):
                            absorption.remove( (row+a-2, col, "r"+str(rows-row+1)) )
                    
        excel_abs.save(filename) # 엑셀 저장
        
        return 0


if __name__ == "__main__":
    pass
    """
    * How to Use?
    from HTML_Table_Excel import Table_Excel
    
    TableExcel = Table_Excel( URL_list <(String)list>, ChromeDriver Path <String>)
    TableExcel.makeExel_abs( Excel File Path <String>, Table Header Color by Hex <String> (Default=#F8E0EC) )
    TableExcel.makeExel_sep( Excel File Path <String> )
    
    * Like this
    url = ["http://www.kweather.co.kr/kma/kma_digital.html",
           "https://www.weather.go.kr/weather/observation/currentweather.jsp"
          ]
    TableExcel = Table_Excel( url, "./driver/chromedriver.exe")
    TableExcel.makeExel_abs( "abs-excel.xlsx", "F8ECE0")
    TableExcel.makeExel_sep( "sep-excel.xlsx")
    """


